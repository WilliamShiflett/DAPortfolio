from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import tkinter as tk
from tkinter import filedialog, messagebox

# Solicit input from the user about which files to use and where to save the output
root = tk.Tk()
root.withdraw()
root.copy_from =  filedialog.askopenfilename(initialdir = "C:/",title = "Select file to copy data from.", filetypes = (("Excel Files","*.xlsx"),("all files","*.*")))
root.copy_to = filedialog.askopenfilename(initialdir = "C:/",title = "Select file to copy data to.", filetypes = (("Excel Files","*.xlsx"),("all files","*.*")))
root.save_location = filedialog.asksaveasfilename(initialdir = "C:/",title = "Input name of file to be saved", defaultextension = ".xlsx", filetypes = (("Excel Files","*.xlsx"),("all files","*.*")))

copy_to_wb = load_workbook(filename=root.copy_to, data_only=True)
copy_from_wb = load_workbook(filename=root.copy_from)

number_of_sheets = len(copy_from_wb.sheetnames)
if number_of_sheets != 1:
	raise Exception("Don't take this the wrong way, but you should only have one worksheet in that workbook! Try picking another?") 

copy_to_ws = copy_to_wb[copy_to_wb.sheetnames[0]]
copy_from_ws = copy_from_wb[copy_from_wb.sheetnames[0]]


colorless_row = []
colorless_rows = []
max_row=copy_from_ws.max_row
max_column=copy_from_ws.max_column



# colorless rows first appear in column D (the 4th column)
# to find colorless rows, we search column D, row by row
for col_cells in copy_from_ws.iter_cols(min_col=4, max_col=4):
	for cell in col_cells:
		if cell.fill.start_color.index == "00000000":
			# when we find a colorless row in column 4, 
			# first, we get its coordinate as a tuple, e.g., ('D',19)
			colorless_cell = coordinate_from_string(cell.coordinate)
			# second, we find the row of the colorless cell
			colorless_row = colorless_cell[1]
			#this next part needs simplification to avoid repeating "row=colorless_row"
			# third, we're grabbing values for four columns from each colorless row
			cell_obj_Paragraph_ReqID = copy_from_ws.cell(row=colorless_row, column=1)
			cell_obj_Private_Implementation = copy_from_ws.cell(row=colorless_row, column=4)
			cell_obj_Status = copy_from_ws.cell(row=colorless_row, column=6)
			cell_obj_Responsible_Entites = copy_from_ws.cell(row=colorless_row, column=10)
			colorless_row = [cell_obj_Paragraph_ReqID.value, cell_obj_Private_Implementation.value, cell_obj_Status.value, cell_obj_Responsible_Entites.value]
			#finally, we use that list to update a list of all colorless rows
			colorless_rows.append(colorless_row)

#now we have a list of all rows that need to be updated in the copy_to_wb
#the next step is to match the Paragraph_ReqID element from the list 
#to the Paragraph/ReqID column in the "copy_to_ws" in the "copy_to_wb"

max_row_1=copy_to_ws.max_row
max_column_1=copy_from_ws.max_column

for i in range(1,max_row_1+1):
	cell_obj_Paragraph_ReqID_1 = copy_to_ws.cell(row=i, column=1)
	for each_colorless_row in colorless_rows:
		if cell_obj_Paragraph_ReqID_1.value == each_colorless_row[0]:
			copy_to_ws.cell(row=i, column=4).value = each_colorless_row[1]
			copy_to_ws.cell(row=i, column=7).value = each_colorless_row[2]
			copy_to_ws.cell(row=i, column=11).value = each_colorless_row[3]
	
copy_to_wb.save(root.save_location)

# Display completion message
tk.messagebox.showinfo('Complete message','Program ran successfully! Check for your new file.')